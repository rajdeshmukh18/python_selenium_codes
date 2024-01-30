import openpyxl
from selenium import webdriver
from tabulate import tabulate

# Function to read data from Excel
# Function to read data from Excel
def read_data_from_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Assuming the data is in a tabular format starting from cell A1
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Modified unpacking to handle a dynamic number of columns
        column1, column2, column3, *remaining_columns = row
        data.append((column1, column2, column3))

    workbook.close()
    return data


# Function to read data from Excel and print in a tabular format
def print_data_from_excel(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Assuming the data is in a tabular format starting from cell A1
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    workbook.close()

    # Print data in a tabular format
    headers = [cell.value for cell in sheet[1]]
    print(tabulate(data, headers=headers, tablefmt="grid"))


# Example usage
file_path = "C:\python-selenium\pythonseleniumProject1\Learning Python\RajDeshmukh_DailyUpdates.xlsx"
sheet_name = "Sheet1"
print_data_from_excel(file_path, sheet_name)
# Read data from Excel

